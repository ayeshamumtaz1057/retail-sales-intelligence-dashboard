"""
export_excel.py
---------------
Builds excel/retail_sales_analysis.xlsx - the Excel deliverable.

Sheets
  README          how to use the workbook
  KPI Dashboard   headline metrics, every cell a live formula over Raw Data
  Raw Data        the 9,993 cleaned transactions (pivot table source)
  Monthly Trend   month-by-month sales and profit
  Category / Region / Top Customers   summary tables driven by SUMIFS

Every number in the summary sheets is a formula referencing 'Raw Data', so the
workbook recalculates if the underlying data is replaced.

Run:  python src/export_excel.py

Note: openpyxl writes formulas without cached values, so a freshly generated
workbook reads back as empty to pandas or any previewer until Excel (or
LibreOffice) opens and recalculates it once. The committed copy in excel/ has
already been recalculated, so it displays correctly on GitHub and on download.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
CLEAN = ROOT / "data" / "processed" / "superstore_features.csv"
OUT = ROOT / "excel" / "retail_sales_analysis.xlsx"

NAVY = "101F4F"
BLUE = "2F6FED"
LIGHT = "EEF1F6"

FONT = "Arial"
H1 = Font(name=FONT, size=14, bold=True, color="FFFFFF")
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="6B7280")

FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_BLUE = PatternFill("solid", fgColor=BLUE)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)

THIN = Side(style="thin", color="D7DBE3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);-'
CUR2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%;(0.0%);-'
NUM = '#,##0;(#,##0);-'


def style_header(ws, row: int, ncols: int, fill=FILL_BLUE) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26


def title_bar(ws, text: str, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font, c.fill = H1, FILL_NAVY
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30


def widths(ws, spec: dict) -> None:
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def build() -> None:
    df = pd.read_csv(CLEAN, parse_dates=["order_date"], dtype={"postal_code": str})
    n = len(df)

    wb = Workbook()

    # ---------------------------------------------------------------- Raw Data
    ws = wb.active
    ws.title = "Raw Data"
    cols = ["order_id", "order_date", "order_year", "order_month_name", "segment",
            "region", "state", "city", "category", "sub_category", "product_name",
            "customer_name", "sales", "quantity", "discount", "profit", "shipping_days"]
    data = df[cols].copy()
    headers = [c.replace("_", " ").title() for c in cols]

    ws.append(headers)
    for row in data.itertuples(index=False):
        ws.append(list(row))
    style_header(ws, 1, len(cols))

    for r in range(2, n + 2):
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=13).number_format = CUR2
        ws.cell(row=r, column=15).number_format = PCT
        ws.cell(row=r, column=16).number_format = CUR2

    # A named Excel table makes this a clean PivotTable source.
    ref = f"A1:{get_column_letter(len(cols))}{n + 1}"
    tbl = Table(displayName="SalesData", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)
    ws.freeze_panes = "A2"
    widths(ws, {"A": 16, "B": 12, "C": 10, "D": 12, "E": 13, "F": 10, "G": 16, "H": 16,
                "I": 16, "J": 14, "K": 42, "L": 20, "M": 12, "N": 9, "O": 10,
                "P": 12, "Q": 12})

    SALES = f"'Raw Data'!$M$2:$M${n+1}"
    PROFIT = f"'Raw Data'!$P$2:$P${n+1}"
    QTY = f"'Raw Data'!$N$2:$N${n+1}"
    CAT = f"'Raw Data'!$I$2:$I${n+1}"
    REG = f"'Raw Data'!$F$2:$F${n+1}"
    SEG = f"'Raw Data'!$E$2:$E${n+1}"
    YEAR = f"'Raw Data'!$C$2:$C${n+1}"
    MON = f"'Raw Data'!$D$2:$D${n+1}"
    CUST = f"'Raw Data'!$L$2:$L${n+1}"

    # ------------------------------------------------------------ KPI Dashboard
    k = wb.create_sheet("KPI Dashboard", 0)
    title_bar(k, "RETAIL SALES INTELLIGENCE  |  Executive Summary", 4)
    widths(k, {"A": 30, "B": 20, "C": 16, "D": 46})

    k["A3"] = "Metric"; k["B3"] = "Value"; k["C3"] = "Format"; k["D3"] = "Definition"
    style_header(k, 3, 4)

    metrics = [
        ("Total Sales", f"=SUM({SALES})", CUR2, "Sum of all order-line revenue"),
        ("Total Profit", f"=SUM({PROFIT})", CUR2, "Sum of all order-line profit"),
        ("Profit Margin", f"=IFERROR(B5/B4,0)", PCT, "Total profit / total sales"),
        ("Total Orders", f"=SUMPRODUCT(1/COUNTIF('Raw Data'!A2:A{n+1},'Raw Data'!A2:A{n+1}))",
         NUM, "Distinct order IDs (rows are order lines, not orders)"),
        ("Total Customers",
         f"=SUMPRODUCT(1/COUNTIF('Raw Data'!L2:L{n+1},'Raw Data'!L2:L{n+1}))",
         NUM, "Distinct customer names"),
        ("Units Sold", f"=SUM({QTY})", NUM, "Total quantity across all lines"),
        ("Avg Order Value", "=IFERROR(B4/B7,0)", CUR2, "Total sales / distinct orders"),
        ("Avg Line Value", f"=IFERROR(B4/{n},0)", CUR2, "Total sales / order lines"),
    ]
    for i, (name, formula, fmt, desc) in enumerate(metrics, start=4):
        k.cell(row=i, column=1, value=name).font = BOLD
        c = k.cell(row=i, column=2, value=formula)
        c.font, c.number_format = BODY, fmt
        k.cell(row=i, column=3, value=fmt.split(";")[0]).font = NOTE
        k.cell(row=i, column=4, value=desc).font = NOTE
        for col in range(1, 5):
            k.cell(row=i, column=col).border = BORDER
            if i % 2 == 0:
                k.cell(row=i, column=col).fill = FILL_LIGHT

    k["A14"] = "Sales by Category"
    k["A14"].font = Font(name=FONT, size=11, bold=True)
    k["A15"] = "Category"; k["B15"] = "Sales"; k["C15"] = "Profit"; k["D15"] = "Margin"
    style_header(k, 15, 4)
    for i, cat in enumerate(["Technology", "Furniture", "Office Supplies"], start=16):
        k.cell(row=i, column=1, value=cat).font = BODY
        k.cell(row=i, column=2, value=f'=SUMIFS({SALES},{CAT},$A{i})').number_format = CUR
        k.cell(row=i, column=3, value=f'=SUMIFS({PROFIT},{CAT},$A{i})').number_format = CUR
        k.cell(row=i, column=4, value=f"=IFERROR(C{i}/B{i},0)").number_format = PCT
        for col in range(1, 5):
            k.cell(row=i, column=col).border = BORDER

    k["A20"] = "Note: every value above is a live formula over the 'Raw Data' sheet."
    k["A20"].font = NOTE
    k["A21"] = "Replace Raw Data with a new extract and the whole workbook recalculates."
    k["A21"].font = NOTE

    # ----------------------------------------------------------- Monthly Trend
    m = wb.create_sheet("Monthly Trend")
    title_bar(m, "Monthly Sales & Profit", 5)
    widths(m, {"A": 10, "B": 12, "C": 16, "D": 16, "E": 12})
    m["A3"] = "Year"; m["B3"] = "Month"; m["C3"] = "Sales"; m["D3"] = "Profit"; m["E3"] = "Margin"
    style_header(m, 3, 5)

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    r = 4
    for yr in sorted(df.order_year.unique()):
        for mo in months:
            m.cell(row=r, column=1, value=int(yr)).number_format = "0"
            m.cell(row=r, column=2, value=mo).font = BODY
            m.cell(row=r, column=3,
                   value=f'=SUMIFS({SALES},{YEAR},$A{r},{MON},$B{r})').number_format = CUR
            m.cell(row=r, column=4,
                   value=f'=SUMIFS({PROFIT},{YEAR},$A{r},{MON},$B{r})').number_format = CUR
            m.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)").number_format = PCT
            for col in range(1, 6):
                m.cell(row=r, column=col).border = BORDER
            r += 1
    m.cell(row=r, column=2, value="Total").font = BOLD
    for col, letter in [(3, "C"), (4, "D")]:
        c = m.cell(row=r, column=col, value=f"=SUM({letter}4:{letter}{r-1})")
        c.font, c.number_format = BOLD, CUR
    m.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)").number_format = PCT
    m.freeze_panes = "A4"

    # ---------------------------------------------------------------- Region
    g = wb.create_sheet("Region Analysis")
    title_bar(g, "Region & Segment Performance", 5)
    widths(g, {"A": 18, "B": 16, "C": 16, "D": 12, "E": 12})
    g["A3"] = "Region"; g["B3"] = "Sales"; g["C3"] = "Profit"
    g["D3"] = "Margin"; g["E3"] = "% of Sales"
    style_header(g, 3, 5)
    for i, reg in enumerate(["West", "East", "Central", "South"], start=4):
        g.cell(row=i, column=1, value=reg).font = BODY
        g.cell(row=i, column=2, value=f'=SUMIFS({SALES},{REG},$A{i})').number_format = CUR
        g.cell(row=i, column=3, value=f'=SUMIFS({PROFIT},{REG},$A{i})').number_format = CUR
        g.cell(row=i, column=4, value=f"=IFERROR(C{i}/B{i},0)").number_format = PCT
        g.cell(row=i, column=5, value=f"=IFERROR(B{i}/$B$8,0)").number_format = PCT
        for col in range(1, 6):
            g.cell(row=i, column=col).border = BORDER
    g.cell(row=8, column=1, value="Total").font = BOLD
    g.cell(row=8, column=2, value="=SUM(B4:B7)").number_format = CUR
    g.cell(row=8, column=3, value="=SUM(C4:C7)").number_format = CUR
    g.cell(row=8, column=4, value="=IFERROR(C8/B8,0)").number_format = PCT
    for col in range(1, 5):
        g.cell(row=8, column=col).font = BOLD

    g["A11"] = "Customer Segment"
    g["A11"].font = Font(name=FONT, size=11, bold=True)
    g["A12"] = "Segment"; g["B12"] = "Sales"; g["C12"] = "Profit"; g["D12"] = "Margin"
    style_header(g, 12, 4)
    for i, seg in enumerate(["Consumer", "Corporate", "Home Office"], start=13):
        g.cell(row=i, column=1, value=seg).font = BODY
        g.cell(row=i, column=2, value=f'=SUMIFS({SALES},{SEG},$A{i})').number_format = CUR
        g.cell(row=i, column=3, value=f'=SUMIFS({PROFIT},{SEG},$A{i})').number_format = CUR
        g.cell(row=i, column=4, value=f"=IFERROR(C{i}/B{i},0)").number_format = PCT
        for col in range(1, 5):
            g.cell(row=i, column=col).border = BORDER

    # -------------------------------------------------------- Top Customers
    t = wb.create_sheet("Top Customers")
    title_bar(t, "Top 20 Customers by Sales", 4)
    widths(t, {"A": 8, "B": 26, "C": 16, "D": 16})
    t["A3"] = "Rank"; t["B3"] = "Customer"; t["C3"] = "Sales"; t["D3"] = "Profit"
    style_header(t, 3, 4)

    # Ranking is done in pandas: Excel's dynamic-array SORT is not portable.
    top = (df.groupby("customer_name")["sales"].sum()
             .sort_values(ascending=False).head(20).index.tolist())
    for i, name in enumerate(top, start=4):
        t.cell(row=i, column=1, value=i - 3).font = BODY
        t.cell(row=i, column=2, value=name).font = BODY
        t.cell(row=i, column=3, value=f'=SUMIFS({SALES},{CUST},$B{i})').number_format = CUR
        t.cell(row=i, column=4, value=f'=SUMIFS({PROFIT},{CUST},$B{i})').number_format = CUR
        for col in range(1, 5):
            t.cell(row=i, column=col).border = BORDER
    t.cell(row=25, column=2,
           value="Ranking computed in Python (src/export_excel.py); values are live SUMIFS."
           ).font = NOTE

    # ---------------------------------------------------------------- README
    rd = wb.create_sheet("README", 0)
    title_bar(rd, "How to use this workbook", 2)
    widths(rd, {"A": 24, "B": 92})
    rows = [
        ("Source", "Sample Superstore, cleaned by src/clean_data.py (9,993 transactions)"),
        ("Coverage", "3 Jan 2015 - 30 Dec 2018, 49 US states, 793 customers"),
        ("", ""),
        ("Raw Data", "The cleaned transactions as an Excel table named 'SalesData'. "
                     "Use it as the source for any PivotTable."),
        ("KPI Dashboard", "Headline metrics. Every value is a live formula, not a pasted number."),
        ("Monthly Trend", "Sales and profit per month via SUMIFS on year and month."),
        ("Region Analysis", "Region and segment breakdown."),
        ("Top Customers", "Top 20 by revenue."),
        ("", ""),
        ("To refresh", "Replace the rows on 'Raw Data' keeping the same column order, "
                       "then press Ctrl+Alt+F9 to recalculate."),
        ("To add a pivot", "Insert > PivotTable > Table/Range: SalesData"),
        ("Distinct counts", "Orders and customers use SUMPRODUCT(1/COUNTIF(...)) because "
                            "Excel has no COUNTDISTINCT function."),
    ]
    for i, (a, b) in enumerate(rows, start=3):
        rd.cell(row=i, column=1, value=a).font = BOLD
        c = rd.cell(row=i, column=2, value=b)
        c.font = BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved -> {OUT.relative_to(ROOT)}  ({n:,} transactions)")


if __name__ == "__main__":
    build()
